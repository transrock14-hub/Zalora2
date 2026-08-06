#!/usr/bin/env python3
"""
Generate a multi-category product catalogue (Excel + CSV) for local review
and Laravel/MySQL-style imports. No product images.
"""

from __future__ import annotations

import csv
import hashlib
import json
import os
import random
import re
import string
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT_DIR = ROOT / "catalog"
XLSX_PATH = OUT_DIR / "zalora-product-catalog.xlsx"
CSV_DIR = OUT_DIR / "laravel-mysql-csv"
SEED = 42
TARGET_PRODUCTS = 3000
MIN_VARIANTS = 10_000

rng = random.Random(SEED)

# ---------------------------------------------------------------------------
# Taxonomy: multi-department (not fashion-only)
# ---------------------------------------------------------------------------

CATEGORY_TREE: dict[str, list[str]] = {
    "Women Fashion": [
        "Dresses", "Tops & Blouses", "Jeans & Denim", "Trousers", "Skirts",
        "Outerwear", "Activewear", "Lingerie", "Swimwear", "Plus Size",
    ],
    "Men Fashion": [
        "Shirts", "T-Shirts & Polos", "Jeans", "Chinos & Trousers", "Jackets",
        "Suits & Blazers", "Activewear", "Underwear", "Swimwear", "Big & Tall",
    ],
    "Kids & Baby": [
        "Boys Clothing", "Girls Clothing", "Baby Essentials", "Kids Footwear",
        "School Uniforms", "Toys & Games", "Kids Accessories",
    ],
    "Shoes": [
        "Women Sneakers", "Women Heels", "Women Boots", "Men Sneakers",
        "Men Formal Shoes", "Men Boots", "Kids Shoes", "Sandals & Flip Flops",
        "Sports Shoes",
    ],
    "Bags & Accessories": [
        "Handbags", "Backpacks", "Wallets", "Belts", "Hats & Caps",
        "Scarves", "Sunglasses", "Watches", "Jewellery", "Luggage",
    ],
    "Beauty & Personal Care": [
        "Skincare", "Makeup", "Hair Care", "Fragrance", "Men Grooming",
        "Bath & Body", "Nail Care", "Oral Care", "Tools & Brushes",
    ],
    "Electronics": [
        "Smartphones & Accessories", "Laptops & Tablets", "Audio & Headphones",
        "Cameras", "Wearables", "Smart Home", "Gaming", "Computer Peripherals",
        "Power Banks & Chargers", "TV & Streaming",
    ],
    "Home & Living": [
        "Furniture", "Bedding", "Bath", "Kitchenware", "Cookware",
        "Home Decor", "Lighting", "Storage & Organisation", "Curtains & Rugs",
        "Cleaning Supplies",
    ],
    "Sports & Outdoors": [
        "Fitness Equipment", "Running", "Team Sports", "Yoga & Pilates",
        "Cycling", "Camping & Hiking", "Swimming", "Sports Nutrition",
        "Outdoor Apparel",
    ],
    "Health & Wellness": [
        "Vitamins & Supplements", "Medical Devices", "First Aid",
        "Massage & Relaxation", "Sleep & Wellness", "Healthcare Essentials",
    ],
    "Food & Grocery": [
        "Pantry Staples", "Snacks & Confectionery", "Beverages",
        "Organic & Health Foods", "Coffee & Tea", "Instant Meals",
        "Baking Supplies",
    ],
    "Pet Supplies": [
        "Dog Food & Treats", "Cat Food & Treats", "Pet Toys",
        "Pet Grooming", "Pet Beds & Housing", "Pet Accessories",
    ],
    "Automotive": [
        "Car Accessories", "Car Care", "Motorcycle Gear",
        "Oils & Fluids", "Tools & Equipment", "Interior Accessories",
    ],
    "Books & Stationery": [
        "Fiction", "Non-Fiction", "Children Books", "Notebooks & Journals",
        "Office Supplies", "Art & Craft Supplies", "Planners & Diaries",
    ],
    "Garden & DIY": [
        "Garden Tools", "Plants & Seeds", "Outdoor Furniture",
        "Power Tools", "Hand Tools", "Paint & Hardware", "BBQ & Outdoor Cooking",
    ],
    "Office & Business": [
        "Desks & Chairs", "Printers & Ink", "Filing & Storage",
        "Presentation Tools", "Business Electronics",
    ],
}

BRANDS = [
    "Nordvik", "Aurelia", "Keystone", "Lumenora", "Harbor & Co", "Solstice",
    "Verdant", "Atlas Peak", "Cinder Lane", "Bluefinch", "Oak & Ember",
    "Polaris Home", "NovaByte", "Silkroad Basics", "Fieldcraft", "UrbanMesh",
    "Canopy Labs", "Driftwood", "Summit Forge", "Paper & Pine", "Voltura",
    "Kelp & Clay", "RoverPet", "Motorlane", "Helix Care", "Brightfolio",
    "Cascade Brew", "TerraForm Tools", "Nightingale", "Coppernest",
]

SUPPLIERS = [
    ("SUP-SG-01", "Apex Wholesale Pte Ltd", "Singapore", "ops@apexwholesale.sg"),
    ("SUP-MY-02", "Peninsula Trade Sdn Bhd", "Malaysia", "sales@peninsulatrade.my"),
    ("SUP-ID-03", "Nusantara Supply Co", "Indonesia", "hello@nusantarasupply.id"),
    ("SUP-TH-04", "Siam Meridian Ltd", "Thailand", "buy@siammeridian.th"),
    ("SUP-CN-05", "Pearl Delta Trading", "China", "export@pearldelta.cn"),
    ("SUP-VN-06", "Saigon Catalog Partners", "Vietnam", "trade@saigoncatalog.vn"),
    ("SUP-HK-07", "Harbour Gate Logistics", "Hong Kong", "desk@harbourgate.hk"),
    ("SUP-PH-08", "Manila Bridge Imports", "Philippines", "orders@manilabridge.ph"),
]

WAREHOUSES = [
    ("WH-SG-EAST", "Singapore East DC", "Singapore", "SUP-SG-01"),
    ("WH-SG-WEST", "Singapore West DC", "Singapore", "SUP-SG-01"),
    ("WH-MY-KL", "Kuala Lumpur Hub", "Malaysia", "SUP-MY-02"),
    ("WH-MY-JB", "Johor Link Warehouse", "Malaysia", "SUP-MY-02"),
    ("WH-ID-JKT", "Jakarta Central FC", "Indonesia", "SUP-ID-03"),
    ("WH-TH-BKK", "Bangkok Fulfilment", "Thailand", "SUP-TH-04"),
    ("WH-CN-SZ", "Shenzhen Export Hub", "China", "SUP-CN-05"),
    ("WH-VN-HCM", "Ho Chi Minh Gateway", "Vietnam", "SUP-VN-06"),
]

COLORS = [
    "Black", "White", "Navy", "Charcoal", "Grey", "Beige", "Ivory", "Brown",
    "Olive", "Forest Green", "Burgundy", "Maroon", "Coral", "Blush Pink",
    "Mustard", "Camel", "Sky Blue", "Teal", "Lavender", "Rust", "Cream",
    "Silver", "Gold", "Rose Gold", "Graphite", "Mint", "Cobalt", "Sand",
]

SIZES_APPAREL = ["XS", "S", "M", "L", "XL", "XXL"]
SIZES_SHOES = ["36", "37", "38", "39", "40", "41", "42", "43", "44", "45"]
SIZES_KIDS = ["2T", "3T", "4", "5", "6", "7", "8", "10", "12"]
SIZES_ONE = ["One Size"]
SIZES_ELECTRONICS = ["Standard", "Pro", "Lite"]
SIZES_HOME = ["Small", "Medium", "Large", "XL"]
SIZES_BEAUTY = ["30ml", "50ml", "100ml", "200ml", "Travel"]
SIZES_FOOD = ["250g", "500g", "1kg", "2kg", "Pack of 6", "Pack of 12"]
SIZES_SPORT = ["S", "M", "L", "XL", "2.5kg", "5kg", "10kg"]

FIRST_NAMES = [
    "Aisha", "Ben", "Chloe", "Danial", "Elena", "Farid", "Grace", "Hassan",
    "Ivy", "Jonas", "Kai", "Lina", "Marcus", "Nora", "Omar", "Priya",
    "Quinn", "Ravi", "Siti", "Tom", "Uma", "Victor", "Wei", "Yasmin", "Zara",
]

REVIEW_TITLES = [
    "Exactly as described", "Solid everyday pick", "Worth the price",
    "Better than expected", "Good build quality", "Arrived quickly",
    "Would buy again", "Decent for the money", "Love this", "Not bad at all",
    "Useful addition", "Reliable so far", "Nice finish", "Happy with it",
]

REVIEW_BODIES = [
    "Packaging was neat and the item matched the listing details.",
    "Used it for two weeks and it holds up well for daily use.",
    "Colour is accurate and sizing feels true based on the chart.",
    "Good value compared with similar products I have tried.",
    "Instructions were clear and setup took only a few minutes.",
    "Feels durable. A couple of small scuffs from shipping but fine overall.",
    "I bought this as a gift and the recipient liked it immediately.",
    "Works as advertised. Battery/performance is satisfactory so far.",
    "Material quality is better than I expected at this price point.",
    "Would appreciate more colour options, but this shade works for me.",
]


def slugify(text: str) -> str:
    s = text.lower().strip()
    s = re.sub(r"[^a-z0-9]+", "-", s)
    return s.strip("-")[:80]


def ean13(prefix12: str) -> str:
    """Build EAN-13 with valid check digit from 12-digit prefix."""
    digits = [int(c) for c in prefix12]
    odd = sum(digits[0::2])
    even = sum(digits[1::2])
    check = (10 - ((odd + even * 3) % 10)) % 10
    return prefix12 + str(check)


def make_barcode(seed_str: str) -> str:
    h = hashlib.sha256(seed_str.encode()).hexdigest()
    # Keep GS1-looking company prefix space (850–889 unused retail block for demo)
    body = "888" + "".join(str(int(h[i], 16) % 10) for i in range(9))
    return ean13(body[:12])


@dataclass
class CatNode:
    id: int
    parent_id: int | None
    name: str
    slug: str
    level: int
    path: str
    department: str


def build_categories() -> list[CatNode]:
    cats: list[CatNode] = []
    cid = 1
    for dept, subs in CATEGORY_TREE.items():
        parent = CatNode(
            id=cid,
            parent_id=None,
            name=dept,
            slug=slugify(dept),
            level=1,
            path=dept,
            department=dept,
        )
        cats.append(parent)
        pid = cid
        cid += 1
        for sub in subs:
            cats.append(
                CatNode(
                    id=cid,
                    parent_id=pid,
                    name=sub,
                    slug=slugify(f"{dept}-{sub}"),
                    level=2,
                    path=f"{dept} > {sub}",
                    department=dept,
                )
            )
            cid += 1
    return cats


def sizes_for(department: str, subcategory: str) -> list[str]:
    d = department.lower()
    s = subcategory.lower()
    if "shoe" in d or "shoe" in s or "sandal" in s or "boot" in s or "heel" in s:
        return SIZES_SHOES
    if "kid" in d or "baby" in d or "boys" in s or "girls" in s:
        if "footwear" in s or "shoe" in s:
            return SIZES_SHOES
        if "clothing" in s or "uniform" in s:
            return SIZES_KIDS
        return SIZES_ONE
    if "beauty" in d or "fragrance" in s or "skincare" in s or "hair" in s:
        return SIZES_BEAUTY
    if "electronic" in d or "office" in d and "electronics" in s:
        return SIZES_ELECTRONICS
    if "food" in d or "grocery" in d or "nutrition" in s:
        return SIZES_FOOD
    if "sport" in d or "fitness" in s or "yoga" in s:
        if "apparel" in s or "running" in s or "swim" in s:
            return SIZES_APPAREL
        return SIZES_SPORT
    if "home" in d or "garden" in d or "furniture" in s:
        return SIZES_HOME
    if "fashion" in d or "bags" in d or "access" in d:
        if "watch" in s or "jewellery" in s or "sunglass" in s or "wallet" in s:
            return SIZES_ONE
        if "bag" in s or "luggage" in s or "backpack" in s:
            return ["Mini", "Medium", "Large"]
        return SIZES_APPAREL
    if "health" in d or "pet" in d or "auto" in d or "book" in d:
        return SIZES_ONE + SIZES_HOME[:2]
    return SIZES_ONE


def colors_for(department: str, subcategory: str) -> list[str]:
    d = department.lower()
    if "food" in d or "book" in d or "health" in d and "device" in subcategory.lower():
        return ["Default", "Assorted"]
    if "electronic" in d:
        return ["Black", "White", "Silver", "Graphite", "Blue"]
    if "beauty" in d:
        return ["Clear", "Pink", "Nude", "Natural", "Black"]
    return COLORS


def product_type_phrases(department: str, subcategory: str) -> list[str]:
    """Concrete product nouns per subcategory — variety beyond clothing."""
    mapping = {
        "Dresses": ["midi wrap dress", "pleated shirt dress", "knit column dress", "linen sundress"],
        "Tops & Blouses": ["silk blouse", "ribbed tank", "poplin shirt", "boat-neck tee"],
        "Jeans & Denim": ["straight-leg jean", "high-rise skinny jean", "wide-leg denim", "cropped flare jean"],
        "Trousers": ["wide-leg trouser", "tailored cigarette pant", "linen drawstring pant"],
        "Skirts": ["A-line midi skirt", "pleated tennis skirt", "denim mini skirt"],
        "Outerwear": ["wool-blend coat", "cropped bomber", "trench coat", "quilted puffer"],
        "Activewear": ["sculpt legging", "mesh panel sports bra", "running short"],
        "Lingerie": ["lace bralette set", "seamless brief pack", "cotton camisole"],
        "Swimwear": ["one-piece swimsuit", "high-waist bikini set", "rash guard"],
        "Plus Size": ["stretch midi dress", "easy-fit blouse", "comfort stretch jean"],
        "Shirts": ["oxford button-down", "linen summer shirt", "twill work shirt"],
        "T-Shirts & Polos": ["heavyweight crew tee", "pique polo", "logo graphic tee"],
        "Jeans": ["slim selvedge jean", "relaxed carpenter jean", "raw denim straight"],
        "Chinos & Trousers": ["stretch chino", "tapered dress pant", "cargo utility pant"],
        "Jackets": ["field jacket", "leather-look biker", "softshell windbreaker"],
        "Suits & Blazers": ["single-breasted blazer", "travel stretch suit", "linen summer blazer"],
        "Underwear": ["modal trunk pack", "ribbed boxer brief", "performance brief"],
        "Big & Tall": ["extended-fit polo", "tall chino", "plus stretch jean"],
        "Boys Clothing": ["graphic jersey set", "chino short set", "school polo pack"],
        "Girls Clothing": ["ruffle dress", "legging set", "denim dungaree"],
        "Baby Essentials": ["organic onesie pack", "muslin swaddle set", "soft sleepsuit"],
        "Kids Footwear": ["light-up sneaker", "school lace shoe", "water sandal"],
        "School Uniforms": ["pleated skirt uniform", "button shirt uniform", "knit vest"],
        "Toys & Games": ["wood building blocks", "strategy board game", "STEM robot kit"],
        "Kids Accessories": ["character backpack", "knit beanie", "fun sock pack"],
        "Women Sneakers": ["chunky platform sneaker", "low-top leather sneaker", "mesh runner"],
        "Women Heels": ["block heel pump", "strap sandal heel", "kitten heel mule"],
        "Women Boots": ["ankle Chelsea boot", "knee-high leather boot", "lug sole boot"],
        "Men Sneakers": ["court sneaker", "retro runner", "slip-on knit sneaker"],
        "Men Formal Shoes": ["cap-toe oxford", "derby shoe", "loafers"],
        "Men Boots": ["chukka boot", "work boot", "desert boot"],
        "Sandals & Flip Flops": ["cushion flip flop", "sport sandal", "slide sandal"],
        "Sports Shoes": ["cross-train shoe", "court shoe", "trail runner"],
        "Handbags": ["structured tote", "crossbody bag", "shoulder bag"],
        "Backpacks": ["commuter backpack", "laptop backpack", "daypack"],
        "Wallets": ["bi-fold wallet", "cardholder", "zip-around wallet"],
        "Belts": ["reversible leather belt", "canvas webbing belt"],
        "Hats & Caps": ["wool fedora", "dad cap", "bucket hat"],
        "Scarves": ["cashmere blend scarf", "printed silk scarf"],
        "Sunglasses": ["aviator sunglasses", "acetate wayfarer", "sport wrap"],
        "Watches": ["automatic wristwatch", "minimal quartz watch", "smart hybrid watch"],
        "Jewellery": ["layered necklace", "hoop earring set", "signet ring"],
        "Luggage": ["spinner cabin case", "checked hard-shell suitcase", "garment duffel"],
        "Skincare": ["vitamin C serum", "ceramides moisturiser", "gel cleanser", "SPF day cream"],
        "Makeup": ["matte lipstick", "liquid foundation", "brow pencil", "eyeshadow palette"],
        "Hair Care": ["repair shampoo", "keratin conditioner", "heat protect spray"],
        "Fragrance": ["eau de parfum", "eau de toilette", "body mist"],
        "Men Grooming": ["beard oil", "face scrub", "electric trimmer kit"],
        "Bath & Body": ["body wash", "exfoliating scrub", "hand cream"],
        "Nail Care": ["gel-effect polish", "nail care kit", "cuticle oil"],
        "Oral Care": ["electric toothbrush", "whitening toothpaste", "floss pack"],
        "Tools & Brushes": ["makeup brush set", "hair dryer", "facial roller"],
        "Smartphones & Accessories": ["phone case", "tempered glass protector", "MagSafe charger", "wireless earbuds case"],
        "Laptops & Tablets": ["14-inch ultraportable laptop", "Android tablet", "laptop sleeve"],
        "Audio & Headphones": ["over-ear headphone", "true wireless earbud", "portable Bluetooth speaker"],
        "Cameras": ["mirrorless camera", "action camera", "webcam"],
        "Wearables": ["fitness tracker", "smartwatch", "HR chest strap"],
        "Smart Home": ["smart plug pack", "Wi-Fi security camera", "smart LED bulb kit"],
        "Gaming": ["wireless controller", "mechanical gaming keyboard", "gaming mouse"],
        "Computer Peripherals": ["ergonomic mouse", "USB-C hub", "wireless keyboard"],
        "Power Banks & Chargers": ["20,000mAh power bank", "GaN wall charger", "car charger"],
        "TV & Streaming": ["streaming stick", "soundbar", "HDMI cable pack"],
        "Furniture": ["accent armchair", "oak side table", "bookshelf unit"],
        "Bedding": ["cotton duvet set", "memory foam pillow", "fitted sheet set"],
        "Bath": ["turkish towel set", "bamboo bath mat", "shower caddy"],
        "Kitchenware": ["ceramic dinner set", "glass food container set", "utensil crock"],
        "Cookware": ["non-stick frying pan", "stainless saucepan", "cast iron skillet"],
        "Home Decor": ["ceramic vase", "framed wall print", "scented candle set"],
        "Lighting": ["desk lamp", "floor lamp", "LED pendant light"],
        "Storage & Organisation": ["modular drawer unit", "closet organiser", "cable box"],
        "Curtains & Rugs": ["blackout curtain panel", "woven area rug", "door mat"],
        "Cleaning Supplies": ["microfibre cloth pack", "steam mop", "vacuum refill kit"],
        "Fitness Equipment": ["adjustable dumbbell", "resistance band set", "yoga mat"],
        "Running": ["hydration belt", "running cap", "reflective armband"],
        "Team Sports": ["football", "basketball", "training cone set"],
        "Yoga & Pilates": ["cork yoga block", "pilates ring", "strap set"],
        "Cycling": ["bike lock", "helmet", "bottle cage"],
        "Camping & Hiking": ["2-person tent", "sleeping bag", "trekking pole set"],
        "Swimming": ["silicone swim cap", "goggles", "kickboard"],
        "Sports Nutrition": ["whey protein", "electrolyte powder", "protein bar box"],
        "Outdoor Apparel": ["rain shell jacket", "hiking pant", "fleece midlayer"],
        "Vitamins & Supplements": ["vitamin D3 softgels", "omega-3 fish oil", "multivitamin"],
        "Medical Devices": ["digital thermometer", "blood pressure monitor", "pulse oximeter"],
        "First Aid": ["first aid kit", "adhesive bandage pack", "antiseptic spray"],
        "Massage & Relaxation": ["percussion massage gun", "foam roller", "heating pad"],
        "Sleep & Wellness": ["weighted blanket", "sleep mask", "white noise machine"],
        "Healthcare Essentials": ["hand sanitiser pack", "face mask box", "disinfectant wipes"],
        "Pantry Staples": ["extra virgin olive oil", "basmati rice", "dried pasta"],
        "Snacks & Confectionery": ["mixed nut tin", "dark chocolate bar", "rice cracker pack"],
        "Beverages": ["sparkling water case", "cold-brew coffee bottle", "green tea bottles"],
        "Organic & Health Foods": ["organic granola", "plant protein powder", "chia seed pouch"],
        "Coffee & Tea": ["single-origin coffee beans", "earl grey tea tin", "matcha powder"],
        "Instant Meals": ["ramen multipack", "ready curry pouch", "oats cup pack"],
        "Baking Supplies": ["cake flour", "vanilla extract", "silicone bakeware set"],
        "Dog Food & Treats": ["adult dry dog food", "dental chew sticks", "training treats"],
        "Cat Food & Treats": ["grain-free cat kibble", "pâté wet food pack", "catnip treats"],
        "Pet Toys": ["rope tug toy", "interactive puzzle toy", "feather wand"],
        "Pet Grooming": ["slicker brush", "pet shampoo", "nail clipper"],
        "Pet Beds & Housing": ["orthopaedic pet bed", "cat tree condo", "travel carrier"],
        "Pet Accessories": ["adjustable leash", "food bowl set", "ID tag"],
        "Car Accessories": ["phone mount", "trunk organiser", "seat gap filler"],
        "Car Care": ["car shampoo kit", "microfibre detailing pack", "tyre shine"],
        "Motorcycle Gear": ["full-face helmet", "riding glove", "tank bag"],
        "Oils & Fluids": ["engine oil 5W-30", "brake fluid", "coolant concentrate"],
        "Tools & Equipment": ["socket wrench set", "jump starter pack", "obd scanner"],
        "Interior Accessories": ["leather seat cover", "air freshener pack", "dashboard mat"],
        "Fiction": ["contemporary novel", "mystery thriller paperback", "literary fiction hardcover"],
        "Non-Fiction": ["business strategy hardcover", "biography paperback", "self-help guide"],
        "Children Books": ["picture storybook", "early reader set", "activity sticker book"],
        "Notebooks & Journals": ["dotted bullet journal", "hardcover notebook", "refillable planner insert"],
        "Office Supplies": ["stapler set", "ballpoint pen pack", "sticky note cube"],
        "Art & Craft Supplies": ["acrylic paint set", "sketch pencil kit", "washi tape pack"],
        "Planners & Diaries": ["undated weekly planner", "desk calendar", "habit tracker"],
        "Garden Tools": ["pruning shears", "hand trowel set", "watering can"],
        "Plants & Seeds": ["herb seed starter kit", "succulent plant set", "tomato seed pack"],
        "Outdoor Furniture": ["patio chair pair", "folding side table", "garden bench cushion"],
        "Power Tools": ["cordless drill kit", "angle grinder", "jigsaw"],
        "Hand Tools": ["carpenter hammer", "screwdriver bit set", "measuring tape"],
        "Paint & Hardware": ["interior matt paint", "roller brush kit", "wall anchor pack"],
        "BBQ & Outdoor Cooking": ["charcoal grill", "BBQ tool set", "smoker chips"],
        "Desks & Chairs": ["standing desk converter", "mesh office chair", "monitor riser"],
        "Printers & Ink": ["inkjet printer", "black ink cartridge", "photo paper pack"],
        "Filing & Storage": ["expanding file folder", "document box", "label maker"],
        "Presentation Tools": ["laser pointer", "portable projector", "whiteboard kit"],
        "Business Electronics": ["conference speakerphone", "document scanner", "label printer"],
    }
    return mapping.get(subcategory, [f"{subcategory.lower()} item"])


def attributes_for(department: str, subcategory: str, product_type: str) -> dict:
    d = department.lower()
    attrs: dict = {
        "department": department,
        "product_type": product_type,
        "origin_country": rng.choice(["SG", "MY", "ID", "TH", "CN", "VN", "KR", "JP", "DE", "IT"]),
        "warranty_months": rng.choice([0, 3, 6, 12, 24]),
    }
    if "fashion" in d or "shoe" in d or "bag" in d:
        attrs.update(
            {
                "material": rng.choice(["Cotton", "Polyester", "Linen", "Leather", "Wool blend", "Nylon", "Denim"]),
                "fit": rng.choice(["Regular", "Slim", "Relaxed", "Oversized"]),
                "care": rng.choice(["Machine wash cold", "Dry clean", "Hand wash", "Wipe clean"]),
                "season": rng.choice(["SS", "AW", "All-season"]),
            }
        )
    elif "beauty" in d:
        attrs.update(
            {
                "skin_type": rng.choice(["All", "Dry", "Oily", "Combination", "Sensitive"]),
                "finish": rng.choice(["Matte", "Dewy", "Natural", "N/A"]),
                "volume_ml": rng.choice([15, 30, 50, 100, 200]),
                "cruelty_free": rng.choice([True, False]),
            }
        )
    elif "electronic" in d or "office" in d:
        attrs.update(
            {
                "connectivity": rng.choice(["Bluetooth 5.3", "USB-C", "Wi-Fi 6", "3.5mm", "Lightning"]),
                "power": rng.choice(["Battery", "Mains", "USB powered", "Rechargeable"]),
                "voltage": rng.choice(["5V", "100-240V", "12V", "N/A"]),
            }
        )
    elif "home" in d or "garden" in d:
        attrs.update(
            {
                "dimensions_cm": f"{rng.randint(10,120)}x{rng.randint(10,80)}x{rng.randint(2,60)}",
                "weight_kg": round(rng.uniform(0.2, 25), 2),
                "assembly_required": rng.choice([True, False]),
                "room": rng.choice(["Living", "Bedroom", "Kitchen", "Bath", "Outdoor", "Office"]),
            }
        )
    elif "sport" in d:
        attrs.update(
            {
                "activity": subcategory,
                "level": rng.choice(["Beginner", "Intermediate", "Advanced"]),
                "weight_kg": round(rng.uniform(0.1, 40), 2),
            }
        )
    elif "food" in d:
        attrs.update(
            {
                "dietary": rng.choice(["None", "Organic", "Vegan", "Halal", "Gluten-free"]),
                "shelf_life_days": rng.randint(30, 720),
                "storage": rng.choice(["Ambient", "Refrigerate", "Cool dry place"]),
            }
        )
    elif "pet" in d:
        attrs.update(
            {
                "pet_type": rng.choice(["Dog", "Cat", "Both", "Small animal"]),
                "life_stage": rng.choice(["Puppy/Kitten", "Adult", "Senior", "All"]),
            }
        )
    elif "auto" in d:
        attrs.update(
            {
                "vehicle_fit": rng.choice(["Universal", "Sedan", "SUV", "Motorcycle"]),
                "material": rng.choice(["ABS plastic", "Leatherette", "Metal", "Rubber"]),
            }
        )
    elif "health" in d:
        attrs.update(
            {
                "form": rng.choice(["Tablet", "Softgel", "Device", "Kit", "Liquid"]),
                "recommended_use": rng.choice(["Daily", "As needed", "Nightly"]),
            }
        )
    elif "book" in d:
        attrs.update(
            {
                "format": rng.choice(["Paperback", "Hardcover", "Set"]),
                "language": rng.choice(["English", "Bilingual EN/MS", "English"]),
                "pages": rng.randint(48, 520),
            }
        )
    else:
        attrs["notes"] = "General merchandise"
    return attrs


def price_for(department: str) -> tuple[float, float, float]:
    """Return retail, compare_at, wholesale."""
    bands = {
        "Women Fashion": (18, 120),
        "Men Fashion": (15, 140),
        "Kids & Baby": (8, 65),
        "Shoes": (25, 180),
        "Bags & Accessories": (12, 220),
        "Beauty & Personal Care": (8, 95),
        "Electronics": (15, 899),
        "Home & Living": (10, 450),
        "Sports & Outdoors": (12, 320),
        "Health & Wellness": (8, 180),
        "Food & Grocery": (3, 55),
        "Pet Supplies": (5, 120),
        "Automotive": (8, 250),
        "Books & Stationery": (4, 65),
        "Garden & DIY": (8, 380),
        "Office & Business": (12, 520),
    }
    lo, hi = bands.get(department, (10, 100))
    retail = round(rng.uniform(lo, hi), 2)
    compare = round(retail * rng.uniform(1.05, 1.35), 2)
    wholesale = round(retail / 1.20, 2)
    return retail, compare, wholesale


def seo_for(name: str, department: str, subcategory: str) -> tuple[str, str]:
    title = f"{name} | {subcategory} | Buy Online"
    if len(title) > 70:
        title = f"{name[:50]} | {subcategory}"[:70]
    meta = (
        f"Shop {name} in {subcategory} ({department}). "
        f"Fast shipping, verified sellers, secure checkout. Order today."
    )[:160]
    return title, meta


HEADER_FILL = PatternFill("solid", fgColor="1A1A1A")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Border(
    left=Side(style="thin", color="DDDDDD"),
    right=Side(style="thin", color="DDDDDD"),
    top=Side(style="thin", color="DDDDDD"),
    bottom=Side(style="thin", color="DDDDDD"),
)


def style_header(ws, cols: int):
    for c in range(1, cols + 1):
        cell = ws.cell(1, c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def autosize(ws, max_width=42):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = 0
        for cell in col[:80]:
            if cell.value is not None:
                width = max(width, min(len(str(cell.value)), max_width))
        ws.column_dimensions[letter].width = max(10, width + 2)


def write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append(row)
    style_header(ws, len(headers))
    autosize(ws)
    return ws


def write_csv(name: str, headers: list[str], rows: list[list]):
    path = CSV_DIR / name
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        w.writerows(rows)
    return path


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    CSV_DIR.mkdir(parents=True, exist_ok=True)

    categories = build_categories()
    leaf_cats = [c for c in categories if c.level == 2]
    assert len(categories) >= 100, f"Need 100+ categories, got {len(categories)}"

    # Allocate products across leaf categories with slight weighting for larger depts
    weights = []
    for c in leaf_cats:
        base = 1.0
        if c.department in ("Electronics", "Home & Living", "Women Fashion", "Beauty & Personal Care"):
            base = 1.35
        if c.department in ("Food & Grocery", "Books & Stationery"):
            base = 0.85
        weights.append(base)
    weight_sum = sum(weights)
    counts = [max(8, int(TARGET_PRODUCTS * (w / weight_sum))) for w in weights]
    # Adjust to exactly TARGET_PRODUCTS
    while sum(counts) > TARGET_PRODUCTS:
        i = counts.index(max(counts))
        counts[i] -= 1
    while sum(counts) < TARGET_PRODUCTS:
        i = counts.index(min(counts))
        counts[i] += 1

    used_names: set[str] = set()
    used_skus: set[str] = set()

    suppliers_rows = []
    for i, (code, name, country, email) in enumerate(SUPPLIERS, start=1):
        suppliers_rows.append([i, code, name, country, email, "ACTIVE"])

    warehouse_rows = []
    for i, (code, name, country, sup_code) in enumerate(WAREHOUSES, start=1):
        warehouse_rows.append([i, code, name, country,sup_code, "ACTIVE"])

    category_rows = [
        [c.id, c.parent_id or "", c.name, c.slug, c.level, c.path, c.department]
        for c in categories
    ]

    product_headers = [
        "id", "sku", "name", "slug", "brand", "department", "category_id", "category_name",
        "category_path", "description", "seo_title", "meta_description", "ean_base",
        "attributes_json", "supplier_code", "warehouse_code", "price", "compare_at_price",
        "wholesale_price", "currency", "rating_avg", "review_count", "stock_total",
        "status", "created_at",
    ]
    variant_headers = [
        "id", "product_id", "product_sku", "variant_sku", "color", "size", "ean", "upc",
        "price", "stock", "warehouse_code", "weight_grams", "status",
    ]
    review_headers = [
        "id", "product_id", "product_sku", "rating", "title", "body", "reviewer_name",
        "verified_purchase", "created_at",
    ]

    products: list[list] = []
    variants: list[list] = []
    reviews: list[list] = []

    pid = 1
    vid = 1
    rid = 1
    base_date = datetime(2024, 1, 1)

    for cat, n_products in zip(leaf_cats, counts):
        phrases = product_type_phrases(cat.department, cat.name)
        size_opts = sizes_for(cat.department, cat.name)
        color_opts = colors_for(cat.department, cat.name)

        for _ in range(n_products):
            brand = rng.choice(BRANDS)
            ptype = rng.choice(phrases)
            # Unique-ish model token
            model = f"{rng.choice(string.ascii_uppercase)}{rng.randint(10,99)}-{rng.randint(100,999)}"
            dlow = cat.department.lower()
            if any(x in dlow for x in ("fashion", "shoe", "bag", "sport")):
                finish_opts = [
                    "", "", "",
                    " — limited edition",
                    " with moisture-wicking lining",
                    " (matte finish)",
                ]
            elif "electronic" in dlow or "office" in dlow:
                finish_opts = ["", "", " compact edition", " pro series", " — limited edition"]
            elif "beauty" in dlow:
                finish_opts = ["", "", " (matte finish)", " travel size", " — limited edition"]
            elif "home" in dlow or "garden" in dlow:
                finish_opts = ["", "", " set", " compact edition"]
            else:
                finish_opts = ["", "", " — limited edition", " value pack"]
            finish = rng.choice(finish_opts)
            name = f"{brand} {ptype.title()} {model}{finish}".strip()
            # Deduplicate names
            attempts = 0
            while name in used_names and attempts < 40:
                model = f"{rng.choice(string.ascii_uppercase)}{rng.randint(10,99)}-{rng.randint(100,999)}"
                name = f"{brand} {ptype.title()} {model}"
                attempts += 1
            if name in used_names:
                name = f"{name}-{pid}"
            used_names.add(name)

            sku = f"ZL{pid:05d}"
            used_skus.add(sku)
            slug = slugify(f"{brand}-{ptype}-{model}-{pid}")
            desc = (
                f"The {name} sits in {cat.path}. "
                f"Built for everyday use with category-appropriate materials and finishing. "
                f"Sold by verified catalogue suppliers; inspect variant options for colour and size."
            )
            seo_title, meta = seo_for(name, cat.department, cat.name)
            attrs = attributes_for(cat.department, cat.name, ptype)
            attrs_json = json.dumps(attrs, ensure_ascii=False)
            price, compare, wholesale = price_for(cat.department)
            supplier = rng.choice(SUPPLIERS)
            warehouse = rng.choice([w for w in WAREHOUSES if w[3] == supplier[0]] or WAREHOUSES)
            ean_base = make_barcode(f"p-{pid}-{sku}")
            rating_avg = round(rng.uniform(3.2, 4.9), 1)
            n_reviews = rng.randint(1, 4)
            created = base_date + timedelta(days=rng.randint(0, 500), hours=rng.randint(0, 23))

            # Variants: need 10k+ total across 3000 products → ~3.4 min average
            n_colors = min(len(color_opts), rng.randint(2, 5))
            n_sizes = min(len(size_opts), rng.randint(2, 5))
            chosen_colors = rng.sample(color_opts, n_colors)
            chosen_sizes = rng.sample(size_opts, n_sizes)
            # Cap per-product variants to keep sheet manageable but hit 10k+
            combo = [(c, s) for c in chosen_colors for s in chosen_sizes]
            rng.shuffle(combo)
            max_v = rng.randint(3, 8)
            combo = combo[:max_v]

            stock_total = 0
            for color, size in combo:
                vsku = f"{sku}-{slugify(color)[:6].upper()}-{slugify(size)[:6].upper()}"
                ean = make_barcode(f"v-{pid}-{vsku}")
                # UPC-A style 12-digit (drop EAN check/prefix for demo column)
                upc = ean[1:13]
                vprice = round(price + rng.uniform(-2, 8), 2)
                if vprice < 1:
                    vprice = price
                stock = rng.randint(0, 120)
                stock_total += stock
                variants.append(
                    [
                        vid, pid, sku, vsku, color, size, ean, upc, vprice, stock,
                        warehouse[0], rng.randint(40, 8500), "ACTIVE",
                    ]
                )
                vid += 1

            for _r in range(n_reviews):
                reviews.append(
                    [
                        rid,
                        pid,
                        sku,
                        rng.randint(3, 5) if rating_avg >= 4 else rng.randint(2, 5),
                        rng.choice(REVIEW_TITLES),
                        rng.choice(REVIEW_BODIES),
                        rng.choice(FIRST_NAMES),
                        rng.choice([1, 1, 1, 0]),
                        (created + timedelta(days=rng.randint(1, 120))).strftime("%Y-%m-%d %H:%M:%S"),
                    ]
                )
                rid += 1

            products.append(
                [
                    pid, sku, name, slug, brand, cat.department, cat.id, cat.name,
                    cat.path, desc, seo_title, meta, ean_base, attrs_json,
                    supplier[0], warehouse[0], price, compare, wholesale, "USD",
                    rating_avg, n_reviews, stock_total, "ACTIVE",
                    created.strftime("%Y-%m-%d %H:%M:%S"),
                ]
            )
            pid += 1

    # Ensure variant count
    if len(variants) < MIN_VARIANTS:
        # Add extra size/color combos to early products
        shortage = MIN_VARIANTS - len(variants)
        print(f"Boosting variants by ~{shortage}...")
        extra_added = 0
        for p in products:
            if extra_added >= shortage:
                break
            pid_x = p[0]
            sku = p[1]
            dept = p[5]
            cat_name = p[7]
            wh = p[15]
            price = float(p[16])
            for color in rng.sample(colors_for(dept, cat_name), k=min(2, len(colors_for(dept, cat_name)))):
                for size in rng.sample(sizes_for(dept, cat_name), k=min(2, len(sizes_for(dept, cat_name)))):
                    if extra_added >= shortage:
                        break
                    vsku = f"{sku}-X{extra_added}-{slugify(color)[:4].upper()}"
                    ean = make_barcode(f"vx-{vid}-{vsku}")
                    variants.append(
                        [
                            vid, pid_x, sku, vsku, color, size, ean, ean[1:13],
                            round(price + 1, 2), rng.randint(5, 80), wh,
                            rng.randint(50, 2000), "ACTIVE",
                        ]
                    )
                    vid += 1
                    extra_added += 1

    # ---- Excel workbook ----
    wb = Workbook()
    # README
    ws = wb.active
    ws.title = "README"
    readme = [
        ["Zalora multi-category product catalogue"],
        ["Generated for local review (no images). Also exported as CSV for Laravel/MySQL import."],
        [""],
        ["Counts"],
        [f"Departments", str(len(CATEGORY_TREE))],
        [f"Categories (incl. parents)", str(len(categories))],
        [f"Leaf subcategories", str(len(leaf_cats))],
        [f"Products", str(len(products))],
        [f"Variants (color/size)", str(len(variants))],
        [f"Reviews", str(len(reviews))],
        [f"Suppliers", str(len(suppliers_rows))],
        [f"Warehouses", str(len(warehouse_rows))],
        [""],
        ["Sheets"],
        ["suppliers", "Supplier master"],
        ["warehouses", "Warehouse / FC master"],
        ["categories", "Department + subcategory tree (parent_id)"],
        ["products", "3000 SKUs with SEO, attributes JSON, ratings"],
        ["variants", "Color/size variants with EAN/UPC"],
        ["reviews", "Customer ratings & review text"],
        ["import_map", "Suggested Laravel / MySQL column mapping"],
        [""],
        ["Notes"],
        ["- No AI/product images included (image columns omitted on purpose)."],
        ["- EAN-13 values include a valid check digit; UPC is derived for demo import."],
        ["- attributes_json varies by department (fashion, beauty, electronics, food, etc.)."],
        ["- CSV copies live in catalog/laravel-mysql-csv/ for artisan imports."],
        [f"- Seed={SEED} for reproducible regenerations."],
        [f"- Generated at {datetime.now().isoformat(timespec='seconds')}"],
    ]
    for row in readme:
        ws.append(row)
    ws["A1"].font = Font(bold=True, size=14)
    autosize(ws, 80)

    write_sheet(
        wb,
        "suppliers",
        ["id", "code", "name", "country", "email", "status"],
        suppliers_rows,
    )
    write_sheet(
        wb,
        "warehouses",
        ["id", "code", "name", "country", "supplier_code", "status"],
        warehouse_rows,
    )
    write_sheet(
        wb,
        "categories",
        ["id", "parent_id", "name", "slug", "level", "path", "department"],
        category_rows,
    )
    write_sheet(wb, "products", product_headers, products)
    write_sheet(wb, "variants", variant_headers, variants)
    write_sheet(wb, "reviews", review_headers, reviews)

    map_rows = [
        ["suppliers", "suppliers", "code unique; link warehouses.supplier_code"],
        ["warehouses", "warehouses", "code unique"],
        ["categories", "categories", "parent_id self-FK; import parents (level=1) first"],
        ["products", "products", "category_id FK; attributes_json → JSON column"],
        ["variants", "product_variants", "product_id FK; unique variant_sku / ean"],
        ["reviews", "product_reviews", "product_id FK"],
        ["", "", "Suggested order: suppliers → warehouses → categories → products → variants → reviews"],
    ]
    write_sheet(
        wb,
        "import_map",
        ["sheet", "suggested_table", "notes"],
        map_rows,
    )

    wb.save(XLSX_PATH)

    # CSV exports
    write_csv("01_suppliers.csv", ["id", "code", "name", "country", "email", "status"], suppliers_rows)
    write_csv("02_warehouses.csv", ["id", "code", "name", "country", "supplier_code", "status"], warehouse_rows)
    write_csv("03_categories.csv", ["id", "parent_id", "name", "slug", "level", "path", "department"], category_rows)
    write_csv("04_products.csv", product_headers, products)
    write_csv("05_variants.csv", variant_headers, variants)
    write_csv("06_reviews.csv", review_headers, reviews)

    print("=== Catalogue ready ===")
    print(f"Excel: {XLSX_PATH}")
    print(f"CSV:   {CSV_DIR}")
    print(f"Categories: {len(categories)} (leaf {len(leaf_cats)}) across {len(CATEGORY_TREE)} departments")
    print(f"Products:   {len(products)}")
    print(f"Variants:   {len(variants)}")
    print(f"Reviews:    {len(reviews)}")
    print(f"Unique names: {len(used_names)}")


if __name__ == "__main__":
    main()
